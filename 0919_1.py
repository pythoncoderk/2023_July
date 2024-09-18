import openpyxl

wb = openpyxl.load_workbook("出社在宅集計表_人事部.xlsx")
print(wb)

ws = wb["4月"]
print(ws)

x = ws.cell(row=2, column=2).value
print(x)

ws.cell(row=5, column=5).value = 1000
wb.save("出社在宅集計表_人事部.xlsx")

wb_matome = openpyxl.load_workbook("出社在宅集計表_まとめ.xlsx")
wb_jinji = openpyxl.load_workbook("出社在宅集計表_人事部.xlsx")

wb_matome = wb_matome["4月"]
wb_jinji = wb_jinji["4月"]

wb_matome.cell(row=2, column=3).value = wb_jinji.cell(row=2, column=2).value
wb.save("出社在宅集計表_まとめ2.xlsx")